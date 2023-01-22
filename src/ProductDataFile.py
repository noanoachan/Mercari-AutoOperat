from typing import List, Tuple
from openpyxl.styles.numbers import builtin_format_code
import os
import openpyxl
import ProductInfo
import Const
import Log
import ExhibitProductInfo

class ReadData:
    """    
     @class     商品情報の読み込み
    """
    # ログクラス
    m_objLog: Log.Log = Log.Log();
    
    
    # 固定値設定クラス
    m_objSettingFile: Const.SettingFile = Const.SettingFile.staticSettingFile();
    
    # 値下げ商品のセル位置
    m_objCellPositionPriceCut: Const.CellPositionPriceCut = Const.CellPositionPriceCut.staticCellPositionPriceCut();
    
    # 出品商品のセル位置
    m_objCellPositionExhibit: Const.CellPositionExhibit = Const.CellPositionExhibit.staticCellPositionExhibit();
    
    
    
    def readPriceCutProduct(self) ->  List[ProductInfo.PriceCutProduct]:
        """
         @detail        値下げ商品リスト取得
         @return        lstPriceCutProduct      値下げ商品情報リスト
        """
        self.m_objLog.LogInfo(f'［シート名：{self.m_objSettingFile.SHEET_PRICE_CUT}］の値下げ商品情報を取得します。');
        
        
        #「値下げ」商品情報リスト
        lstPriceCutProduct: List[ProductInfo.PriceCutProduct] = []
        
        try:
            # 該当の「商品情報」ファイルを取得
            strFilePath: str = os.path.normpath(os.path.join(self.m_objSettingFile.EXE_BASE_FILE_PATH, f'..\\doc\\{self.m_objSettingFile.FILE_PRODUCT_LIST}'));
            xlBook = openpyxl.load_workbook(strFilePath);
            xlSheet = xlBook[self.m_objSettingFile.SHEET_PRICE_CUT];
            
            # 商品数カウンター
            nProductCount: int = 1;
            try:
                for row in xlSheet.iter_rows(min_row=2):
                    # セルの値をリストとして取得
                    xlValues = [cell.value for cell in row];
                    
                    # 商品IDが空なら次へ
                    if xlValues[self.m_objCellPositionPriceCut.PRODUCT_ID] is None:
                        continue;
                    
                    strProductName: str = xlValues[self.m_objCellPositionPriceCut.PRODUCT_NAME];      # 商品名
                    strProductId: str   = xlValues[self.m_objCellPositionPriceCut.PRODUCT_ID];        # 商品ID
                    nCheapestPrice: int = xlValues[self.m_objCellPositionPriceCut.CHEAPES_PRICE];     # 最安値
                    
                    # 値下げ商品情報
                    objPriceCutInfo: ProductInfo.PriceCutProduct = ProductInfo.PriceCutProduct(strProductName, strProductId);
                    
                    # 最安値が「空白」でなければ
                    if nCheapestPrice is not None:
                        objPriceCutInfo.setCheapestPrice(nCheapestPrice);
                        
                    # 数値以外であれば警告を出し「0円」で設定
                    else:
                        objPriceCutInfo.setCheapestPrice(0);
                        self.m_objLog.LogWarning('最安値が「空白」または「文字列」で設定されています');
                        self.m_objLog.LogWarning(f'最安値を「0円」で設定し直しました。商品ID［{strProductId}］');
                    
                    # 商品情報の追加
                    lstPriceCutProduct.append(objPriceCutInfo);
                    
                    self.m_objLog.LogInfo(f'{nProductCount}つ目の商品情報の取得に成功しました。商品ID［{strProductId}］');
                    nProductCount += 1;     # 商品番号のインクリメント
                    
            except Exception as e:
                self.m_objLog.LogError(f'{nProductCount}つ目の商品情報の取得に失敗しました');
                nProductCount += 1;         # 商品番号のインクリメント
                pass;
            
            # 上書き保存
            xlBook.save(strFilePath);
            # 閉じる
            xlBook.close();
            
        except FileNotFoundError as e:
            self.m_objLog.LogCritical(f'{self.m_objSettingFile.FILE_PRODUCT_LIST}の取得に失敗しました');
            raise;
            

        return lstPriceCutProduct;
    
    
    
    def readExhibitProduct(self) ->  List[ExhibitProductInfo.ExhibitProductInfo]:
        """
         @detail        出品商品取得
         @return        lstExhibitProduct   出品商品情報リスト
         @note          ProductList.xlsx(シート名：template, data, master)内の「出品商品」リスト取得
        """
        self.m_objLog.LogInfo(f'{self.m_objSettingFile.FILE_PRODUCT_LIST}内の商品情報を取得します');
        
        # 出品商品のリスト
        lstExhibitProductInfo: List[ExhibitProductInfo.ExhibitProductInfo] = []
        
        try:
            
            # 出品商品情報が入力されている Excelブックを開く
            strExcelBookPath: str = os.path.normpath(os.path.join(self.m_objSettingFile.EXE_BASE_FILE_PATH, f'..\\doc\\{self.m_objSettingFile.FILE_PRODUCT_LIST}'));
            xlBook = openpyxl.load_workbook(strExcelBookPath);
            
            #########################################################
            # シート名：「master」の情報取得
            #########################################################
            self.m_objLog.LogInfo(f'［シート名：{self.m_objSettingFile.SHEET_MASTER}］を取得します');
            
            ExhibitProductCommon: ExhibitProductInfo.ExhibitProductCommon;
            try:
                xlSheetMaster = xlBook[self.m_objSettingFile.SHEET_MASTER];
                
                xlValue_ShippingCharge = xlSheetMaster[self.m_objCellPositionExhibit.SHUPPING_CHARGE].value;    # 配送料の負担
                if xlValue_ShippingCharge == None:
                            xlValue_ShippingCharge = '送料込み（出品者負担）';      #? 記載がなかった時のデフォルト値
                            self.m_objLog.LogInfo(f'「配送料の負担」項目が未設定のためデフォルト値で設定し直しました。配送料の負担［{xlValue_ShippingCharge}］');
                            
                xlValue_ShippingMethod = xlSheetMaster[self.m_objCellPositionExhibit.SHUPPING_METHOD].value;    # 配送の方法
                if xlValue_ShippingMethod == None:
                            xlValue_ShippingMethod = '未定';                      #? 記載がなかった時のデフォルト値
                            self.m_objLog.LogInfo(f'「配送の方法」項目が未設定のためデフォルト値で設定し直しました。配送料の負担［{xlValue_ShippingMethod}］');
                            
                xlValue_ShippingArea = xlSheetMaster[self.m_objCellPositionExhibit.SHUPPING_AREA].value;        # 配送元の地域
                if xlValue_ShippingArea == None:
                            xlValue_ShippingArea = '東京都'                       #? 記載がなかった時のデフォルト値
                            self.m_objLog.LogInfo(f'「配送元の地域」項目が未設定のためデフォルト値で設定し直しました。配送料の負担［{xlValue_ShippingArea}］');
                            
                xlValue_ShippingDays = xlSheetMaster[self.m_objCellPositionExhibit.SHUPPING_DAYS].value;        # 発送までの日数
                if xlValue_ShippingDays == None:
                            xlValue_ShippingDays = '4~7日で発送';                 #? 記載がなかった時のデフォルト値
                            self.m_objLog.LogInfo(f'「発送までの日数」項目が未設定のためデフォルト値で設定し直しました。配送料の負担［{xlValue_ShippingDays}］');
                
                
                # 出品商品情報クラス
                ExhibitProductCommon = ExhibitProductInfo.ExhibitProductCommon(
                    xlValue_ShippingCharge,
                    xlValue_ShippingMethod,
                    xlValue_ShippingArea,
                    xlValue_ShippingDays);

                
            except Exception as e:
                self.m_objLog.LogError(f'［シート名：{self.m_objSettingFile.SHEET_MASTER}］が存在しません');
                
                # デフォルト値に設定
                ExhibitProductCommon = ExhibitProductInfo.ExhibitProductCommon(
                    '送料込み（出品者負担）',
                    '未定',
                    '東京都',
                    '4~7日で発送'
                );
                
                self.m_objLog.LogInfo('発送に関する情報をデフォル値で設定し続行します');
                pass;
            
            
            #########################################################
            # シート名：「template」の情報取得
            #########################################################
            self.m_objLog.LogInfo(f'［シート名：{self.m_objSettingFile.SHEET_TEMPLATE}］を取得します');
            xlSheetTemplate = xlBook[self.m_objSettingFile.SHEET_TEMPLATE];
            xlValueExplanation  = xlSheetTemplate["B5"];       # 商品説明（テンプレート）
            if xlValueExplanation == '':
                self.m_objLog.LogWarning(f'シート名「{self.m_objSettingFile.SHEET_TEMPLATE}」商品説明の雛形が【空白】のため商品説明の入力欄は空白で設定します');
            
            
            #########################################################
            # シート名：「data」の情報取得
            #########################################################
            self.m_objLog.LogInfo(f'［シート名：{self.m_objSettingFile.SHEET_DATA}］を取得します');
                    
            # インクリメント用変数
            nProductCount           : int = 0;          # 全体商品数
            nProductSuccessCount    : int = 0;          # 取得成功商品数
            nProductSkipCount       : int = 0;          # 取得スキップ商品数
            nProductErrorCount      : int = 0;          # 取得失敗商品数
            
            xlSheetData = xlBook[self.m_objSettingFile.SHEET_DATA];
            for row in xlSheetData.iter_rows(min_row=self.m_objCellPositionExhibit.GET_DATA_FIRST_ROW):
                
                try:    
                    # 商品(全体)数のインクリメント
                    nProductCount += 1;
                    # 商品説明の雛形
                    strExplanation: str = xlValueExplanation.value;

                    # セルの値をリストとして取得
                    xlValues: List = [cell.value for cell in row];
                    
                    # 商品No.が空なら次へ
                    if xlValues[self.m_objCellPositionExhibit.PRODUCT_NAME] is None:
                        # 商品(スキップ)数のインクリメント
                        self.m_objLog.LogWarning(f'商品No.[{nProductCount}]の商品が【空白】のためスキップします［シート名：{self.m_objSettingFile.SHEET_DATA}］');
                        nProductSkipCount += 1;
                        continue;
                    
                    # 商品が既に出品済みであれば次へ（R列に"〇"が付いていれば）
                    if xlValues[self.m_objCellPositionExhibit.ALREADY_PRODUCT] == '〇':
                        # 商品(スキップ)数のインクリメント
                        self.m_objLog.LogWarning(f'商品No.[{nProductCount}]の商品が【既に出品中】のためスキップします［シート名：{self.m_objSettingFile.SHEET_DATA}］');
                        nProductSkipCount += 1;
                        continue;
                    

                    # 商品番号
                    strProductNo                : str           = xlValues[self.m_objCellPositionExhibit.PRODUCT_NO];

                    # 商品写真のフォルダ名
                    strProductPhotoFolderName   : str           = xlValues[self.m_objCellPositionExhibit.PRODUCT_PHOTO_FOLDER_PATH];

                    # カテゴリー_1
                    strCategory_1               : str           = xlValues[self.m_objCellPositionExhibit.CATEGORY_1];
                    if strCategory_1            == None:
                        strCategory_1                           = 'その他';                         #? 記載がなかった時のデフォルト値

                    # カテゴリー_2
                    strCategory_2               : str           = xlValues[self.m_objCellPositionExhibit.CATEGORY_2];
                    if strCategory_2            == None:
                        strCategory_2                           = 'その他';                         #? 記載がなかった時のデフォルト値

                    # カテゴリー_3
                    strCategory_3               : str           = xlValues[self.m_objCellPositionExhibit.CATEGORY_3];

                    # ブランド_1
                    strBrand_1                  : str           = xlValues[self.m_objCellPositionExhibit.BRAND_1];

                    # 商品名
                    strProductName              : str           = xlValues[self.m_objCellPositionExhibit.PRODUCT_NAME];
                    if strProductName           == None:
                        strProductName                          = '商品名記入なし';                  #? 記載がなかった時のデフォルト値

                    # 商品状態
                    strCommodityCondition       : str           = xlValues[self.m_objCellPositionExhibit.COMMODITY_CONDITION];
                    if strCommodityCondition    == None:
                        strCommodityCondition                   = '新品、未使用';                   #? 記載がなかった場合のデフォルト値

                    # 出品開始額
                    nStartingPrice              : int           = xlValues[self.m_objCellPositionExhibit.STARTING_PRICE];
                    if nStartingPrice           == None:
                        nStartingPrice                          = 999999;

                    # 商品状態_A
                    strProductDetailsA          : str           = str(xlValues[self.m_objCellPositionExhibit.PRODUCT_DETAILS_A]);
                    if strProductDetailsA       == None:
                        strProductDetailsA                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_B
                    strProductDetailsB          : str           = str(xlValues[self.m_objCellPositionExhibit.PRODUCT_DETAILS_B]);
                    if strProductDetailsB       == None:
                        strProductDetailsB                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_C
                    strProductDetailsC          : str           = str(xlValues[self.m_objCellPositionExhibit.PRODUCT_DETAILS_C]);
                    if strProductDetailsC       == None:
                        strProductDetailsC                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_D
                    strProductDetailsD          : str           = str(xlValues[self.m_objCellPositionExhibit.PRODUCT_DETAILS_D]);
                    if strProductDetailsD       == None:
                        strProductDetailsD                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_E
                    strProductDetailsE          : str           = str(xlValues[self.m_objCellPositionExhibit.PRODUCT_DETAILS_E]);
                    if strProductDetailsE       == None:
                        strProductDetailsE                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_F
                    strProductDetailsF          : str           = str(xlValues[self.m_objCellPositionExhibit.PRODUCT_DETAILS_F]);
                    if strProductDetailsF       == None:
                        strProductDetailsF                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_G
                    strProductDetailsG          : str           = str(xlValues[self.m_objCellPositionExhibit.PRODUCT_DETAILS_G]);
                    if strProductDetailsG       == None:
                        strProductDetailsG                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_H
                    strProductDetailsH          : str           = str(xlValues[self.m_objCellPositionExhibit.PRODUCT_DETAILS_H]);
                    if strProductDetailsH       == None:
                        strProductDetailsH                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_I
                    strProductDetailsI          : str           = str(xlValues[self.m_objCellPositionExhibit.PRODUCT_DETAILS_I]);
                    if strProductDetailsI       == None:
                        strProductDetailsI                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_J
                    strProductDetailsJ          : str           = str(xlValues[self.m_objCellPositionExhibit.PRODUCT_DETAILS_J]);
                    if strProductDetailsJ       == None:
                        strProductDetailsJ                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_K
                    strProductDetailsK          : str           = str(xlValues[self.m_objCellPositionExhibit.PRODUCT_DETAILS_K]);
                    if strProductDetailsK       == None:
                        strProductDetailsK                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_L
                    strProductDetailsL          : str           = str(xlValues[self.m_objCellPositionExhibit.PRODUCT_DETAILS_L]);
                    if strProductDetailsL       == None:
                        strProductDetailsL                      = ' ';                               #? 記載がなかった場合のデフォルト値
                        
                    
                    
                    #############################################################################
                    # 設定されたデータ(商品写真パス)から商品写真の取得
                    #############################################################################
                    # 出品商品写真までのパスからフォルダ内全ての写真ファイル名（絶対パス）を取得
                    
                    # 出品商品の写真が格納されている親までの絶対パス
                    strToPhotoFolderPath: str = os.path.normpath(os.path.join(self.m_objSettingFile.EXE_BASE_FILE_PATH, f'..\\doc\\{self.m_objSettingFile.FOLDER_PRODUCT_PHOTO}'));
                    # 商品写真の指定フォルダを親までの絶対パスと結合
                    strToPhotoFolderPath += f'\\{strProductPhotoFolderName}';
                
                    # フォルダ内全ての商品写真のファイル名を絶対パスで追加      #! (確認用)
                    lstProductPhotoFileName: List = [];
                    for strPhotoFileName in os.listdir(strToPhotoFolderPath):
                        lstProductPhotoFileName.append(f'{strToPhotoFolderPath}\\{strPhotoFileName}');
                    self.m_objLog.LogDebug(f'商品No.[{strProductNo}]内、全てのファイル名\n{lstProductPhotoFileName}');       # フォルダ内に存在する商品写真ファイル名を全て表示
                    
                    
                    if len(lstProductPhotoFileName) < 1:
                        self.m_objLog.LogWarning(f'商品No.[{strProductNo}] の商品写真が 1つも見つかりませんでした');
                    elif len(lstProductPhotoFileName) >= 10:
                        self.m_objLog.LogWarning(f'商品No.[{strProductNo}]の商品写真が 10個(制限枚数)を超えているため 11個目以降の写真は挿入されません');
                        
                        # 挿入されない商品の警告
                        for strProductPhotoFileName in lstProductPhotoFileName[10::-1]:
                            self.m_objLog.LogWarning(f'挿入されない商品写真［{strProductPhotoFileName}］');
                            lstProductPhotoFileName.remove(strProductPhotoFileName);
                            
                        self.m_objLog.LogDebug(f'挿入される写真は以下です［商品No.[{strProductNo}]］\n{lstProductPhotoFileName}');
                        
                    
                    #############################################################################
                    # 設定されたデータ(商品説明の各設定値)から商品説明のテンプレートを作成(整形)
                    #############################################################################
                    # 商品ごとの設定値を「商品説明」の雛形に文字列を置換し整形
                    strExplanation = strExplanation.replace("$A$", strProductDetailsA);
                    strExplanation = strExplanation.replace("$B$", strProductDetailsB);
                    strExplanation = strExplanation.replace("$C$", strProductDetailsC);
                    strExplanation = strExplanation.replace("$D$", strProductDetailsD);
                    strExplanation = strExplanation.replace("$E$", strProductDetailsE);
                    strExplanation = strExplanation.replace("$F$", strProductDetailsF);
                    strExplanation = strExplanation.replace("$G$", strProductDetailsG);
                    strExplanation = strExplanation.replace("$H$", strProductDetailsH);
                    strExplanation = strExplanation.replace("$I$", strProductDetailsI);
                    strExplanation = strExplanation.replace("$J$", strProductDetailsJ);
                    strExplanation = strExplanation.replace("$K$", strProductDetailsK);
                    strExplanation = strExplanation.replace("$L$", strProductDetailsL);
                    self.m_objLog.LogDebug(f'商品No.[{strProductNo}]の商品説明\n{strExplanation}');
                    
                    
                    
                    # 出品商品情報（シート名：data）登録
                    objExhibitProductInfo: ExhibitProductInfo.ExhibitProductInfo = ExhibitProductInfo.ExhibitProductInfo(
                        strProductNo,                   # 商品番号
                        strToPhotoFolderPath,           # 商品写真フォルダまでの絶対パス
                        strCategory_1,                  # カテゴリ_1
                        strCategory_2,                  # カテゴリ_2
                        strCategory_3,                  # カテゴリ_3
                        strBrand_1,                     # ブランド_1
                        strProductName,                 # 商品名
                        strCommodityCondition,          # 商品状態
                        nStartingPrice,                 # 出品価格
                        strExplanation                  # 商品説明
                        );
                    
                    
                    # 共通設定部（シート名：master）登録
                    objExhibitProductInfo.setExhibitProductCommon(ExhibitProductCommon);
                    
                    
                    #############################################################################
                    # 出品商品情報を追加
                    #############################################################################
                    lstExhibitProductInfo.append(objExhibitProductInfo);
                    
                    nProductSuccessCount += 1;      # 商品(成功)数のインクリメント
                    self.m_objLog.LogInfo(f'商品No.[{objExhibitProductInfo.getProductNo()}]の商品情報の取得に成功しました');
                    
                except Exception as e:
                    self.m_objLog.LogError(f'商品No.[{nProductCount}]の商品情報の取得に失敗しました');
                    nProductErrorCount += 1;        # 商品(エラー)数のインクリメント
                    continue;
                
            # 商品数読み込み結果の出力
            self.m_objLog.LogInfo(f'商品全体数：{nProductCount}  成功：{nProductSuccessCount}  スキップ：{nProductSkipCount}  失敗：{nProductErrorCount}');
            
            # 閉じる
            xlBook.close();
            
        except FileNotFoundError as e:
            self.m_objLog.LogCritical(f'{self.m_objSettingFile.FILE_PRODUCT_LIST}の取得に失敗しました');
            return lstExhibitProductInfo;
        
        return lstExhibitProductInfo;
            
                        
class WriteData:
    """
     @class     書き込みクラス
    """
    # ログクラス
    m_objLog: Log.Log = Log.Log();


    # 固定値設定クラス
    m_objSettingFile: Const.SettingFile = Const.SettingFile.staticSettingFile();
    
    # 出品商品のセル位置
    m_objCellPositionExhibit = Const.CellPositionExhibit();
    
       

    def writeGetProductInfo(self, lstWriteProductInfo: List) ->  None:
        """
         @details       値下げ商品取得
         @param[in]     lstWriteProductInfo     値下げ商品情報
        """
        
        objCellData = Const.CellPositionPriceCut();
        
        # データファイルに記録されてある商品情報と出品中の商品情報に差異がない場合、追加記録せず終了
        if len(lstWriteProductInfo) == 0:
            self.m_objLog.LogInfo('データファイルと出品中の商品情報に差異がないため追加記録は行いません');
            return;
        
        try:
            strFilePath     = os.path.normpath(os.path.join(self.m_objSettingFile.EXE_BASE_FILE_PATH, f'..\\doc\\{self.m_objSettingFile.FILE_PRODUCT_LIST}'));
            xlBook          = openpyxl.load_workbook(strFilePath);
            xlSheet         = xlBook[self.m_objSettingFile.SHEET_PRICE_CUT];

            xlMaxRow        : int   = xlSheet.max_row;      # 全範囲での最終行
            xlLastRow       : int   = 1;                    # 必要な商品情報が記載されている範囲での最終行
            nProductCount   : int   = 1;                    # 商品数のインクリメント変数

            # 最終行から逆ループ
            for row in reversed(range(1, xlMaxRow)):
                # 指定列の行が空でなくなるまで遡り最終的な指定列の最終行を探す (※指定列：商品ID)
                if xlSheet.cell(row=row, column=objCellData.PRODUCT_ID + 1).value != None:
                    # 指定列の最終行を取得
                    xlLastRow = row + 1;
                    break;
                    
            try:
                # 最終行から新規商品情報を記載
                for objWriteProductInfo in lstWriteProductInfo:
                    xlSheet.cell(xlLastRow, objCellData.PRODUCT_NAME + 1, value=objWriteProductInfo.getProductName());      # 商品名
                    xlSheet.cell(xlLastRow, objCellData.PRODUCT_ID + 1, value=objWriteProductInfo.getProductId());          # 商品ID
                    
                    self.m_objLog.LogInfo(f'{nProductCount}つ目の商品情報を記録しました。商品ID［{objWriteProductInfo.getProductId()}］');
                    xlLastRow       += 1;       # 書き込み行のインクリメント
                    nProductCount   += 1;       # 商品番号のインクリメント
                    
            except Exception as e:
                nProductCount += 1;     # 商品番号のインクリメント
                self.m_objLog.LogError(f'{nProductCount}つ目の商品情報の記録に失敗しました');
                pass;
                
                
            # 上書き保存
            xlBook.save(strFilePath);
            # 閉じる
            xlBook.close();
            
        except Exception as e:
            self.m_objLog.LogCritical(f'{self.m_objSettingFile.FILE_PRODUCT_LIST}の取得に失敗しました');
        
        
        
    def writePriceCutProductStatus(self, lstWriteProductStatus: List) ->  None:
        """
         @details       商品を値下げ後、その時に取得した商品状態(公開停止、削除、売り切れ)を記録
         @param[in]     lstWriteProductStatus   商品状態リスト（公開停止、削除、売り切れ）
        """
        objCellData = Const.CellPositionPriceCut();
        
        # データファイルの情報と値下げを行った際に得た商品状態に変更がない場合、追加記録せず終了
        if len(lstWriteProductStatus) == 0:
            self.m_objLog.LogInfo('商品状態の新たな変更が確認できないため追加記録は行いません');
            return;
    
        try:
            strFilePath = os.path.normpath(os.path.join(self.m_objSettingFile.EXE_BASE_FILE_PATH, f'..\\doc\\{self.m_objSettingFile.FILE_PRODUCT_LIST}'));
            xlBook = openpyxl.load_workbook(strFilePath);
            xlSheet = xlBook.worksheets[0];

            try:
                # 範囲データを順次処理
                for tupleTargetCellData in xlSheet.iter_cols(min_row=2, min_col=objCellData.PRODUCT_ID + 1, max_col=objCellData.PRODUCT_ID + 1):
                    # タプルで返ってきた商品IDのデータを分析
                    for xlCellData in tupleTargetCellData:
                        
                        # 商品IDの取得
                        strCellData = xlCellData.value;
                        
                        # 商品状態「公開停止、削除、売り切れ」のいずれかに該当する商品IDが存在するか範囲データ内を全検索
                        for objWriteProductStatus in lstWriteProductStatus:
                            
                            #「公開停止、削除、売り切れ」いずれかに該当がある商品IDがデータ内に存在するか否か
                            if strCellData == objWriteProductStatus.getProductId():
                                
                                # 「備考」セル位置を取得
                                xlCellData.offset(0, 2);
                                
                                if objWriteProductStatus.getProductStopPublishing():
                                    xlCellData.offset(0, 2).value = '公開停止';
                                    self.m_objLog.LogInfo(f'商品ID：{objWriteProductStatus.getProductId()} の備考欄に「公開停止」を記録しました');
                                    
                                if objWriteProductStatus.getProductDelete():
                                    xlCellData.offset(0, 2).value = '削除';
                                    self.m_objLog.LogInfo(f'商品ID：{objWriteProductStatus.getProductId()} の備考欄に「削除」を記録しました');
                                    
                                if objWriteProductStatus.getProductSoldOut():
                                    xlCellData.offset(0, 2).value = '売り切れ';
                                    self.m_objLog.LogInfo(f'商品ID：{objWriteProductStatus.getProductId()} の備考欄に「売り切れ」を記録しました');
                                    
            except Exception as e:
                self.m_objLog.LogError(f'商品状態の更新に失敗しました');
                pass;
                
            # 上書き保存
            xlBook.save(strFilePath);
            # 閉じる
            xlBook.close();
            self.m_objLog.LogInfo('全商品状態の更新が完了しました');
                    
        except FileNotFoundError as e:
            self.m_objLog.LogCritical(f'{self.m_objSettingFile.FILE_PRODUCT_LIST}の取得に失敗しました');
            
            

    def writeExhibitProductStatus(self, lstExhibitProductInfo: List) ->  None:
        """
         @details       出品商品を公開できたか否かの結果を記録
         @param[in]     lstExhibitProductInfo   出品に「成功」した商品リスト
        """
        # 商品出品数が「0」の場合は追加記録せず終了
        if len(lstExhibitProductInfo) == 0:
            self.m_objLog.LogInfo('出品した商品数が「0」なため記録は行いません');
            return;
    
        try:
            strFilePath = os.path.normpath(os.path.join(self.m_objSettingFile.EXE_BASE_FILE_PATH, f'..\\doc\\{self.m_objSettingFile.FILE_PRODUCT_LIST}'));
            xlBook = openpyxl.load_workbook(strFilePath);
            xlSheet = xlBook[self.m_objSettingFile.SHEET_DATA];

            try:
                # 範囲データを順次処理
                for tupleTargetCellData in xlSheet.iter_cols(min_row=self.m_objCellPositionExhibit.GET_DATA_FIRST_ROW, min_col=self.m_objCellPositionExhibit.PRODUCT_NO + 1, max_col=self.m_objCellPositionExhibit.PRODUCT_NO + 1):
                    # タプルで返ってきた商品IDのデータを分析
                    for xlCellData in tupleTargetCellData:
                        
                        # 商品No.の取得
                        strCellData = xlCellData.value;
                        
                        # 出品が「成功」した商品No.を順に取り出し比較
                        for objExhibitProductInfo in lstExhibitProductInfo:
                            
                            # 同「商品No.」に"〇"を付与
                            if strCellData == objExhibitProductInfo.getProductNo():
                                
                                # 「出品有無」セル位置を設定 (※対象のセル位置(商品No.)からの相対位置)
                                xlCellData.offset(0, self.m_objCellPositionExhibit.ALREADY_PRODUCT - 1).value = '〇';
                                self.m_objLog.LogInfo(f'商品No.[{objExhibitProductInfo.getProductNo()}]に出品完了マークを付与しました');
                                    
            except Exception as e:
                self.m_objLog.LogError(f'商品状態の更新に失敗しました');
                pass;
                
            # 上書き保存
            xlBook.save(strFilePath);
            # 閉じる
            xlBook.close();
            self.m_objLog.LogInfo('全商品状態の更新が完了しました');
                    
        except FileNotFoundError as e:
            self.m_objLog.LogCritical(f'{self.m_objSettingFile.FILE_PRODUCT_LIST}の取得に失敗しました');



    def writeResetLowestPrice(self) ->  None:
        """
        @ details       再出品商品の最安値設定
        """
        try:
            strFilePath = os.path.normpath(os.path.join(self.m_objSettingFile.EXE_BASE_FILE_PATH, f'..\\doc\\{self.m_objSettingFile.FILE_PRODUCT_LIST}'));
            xlBook      = openpyxl.load_workbook(strFilePath);
            xlSheet     = xlBook[self.m_objSettingFile.SHEET_PRICE_CUT];
            
            try:
                # [ターゲット商品] 指定行~指定列 範囲内のセル値をタプルで取得し 1列単位で処理
                for cols in xlSheet.iter_cols(min_row=2, min_col=1, max_col=1):
                    # [ターゲット商品] 列から各セル値を取り出し処理 (※iter_cols()で範囲を指定しているため「商品名」の 1つしか存在しない)
                    for cellTarget in cols:
                        # [比較商品] 指定行~指定列 範囲内のセル値をタプルで取得し 1列単位で処理
                        for cols in xlSheet.iter_cols(min_row=2, min_col=1, max_col=1):
                            # [比較商品] 列から各セル値を取り出し処理 (※iter_cols()で範囲を指定しているため「商品名」の 1つしか存在しない)
                            for cellComparison in cols:
                                
                                #「ターゲット商品」の「行番号」が「比較商品」の「行番号」より若い場合はスキップ
                                if cellTarget.row >= cellComparison.row:
                                    continue;
                                
                                #「商品名」が同じ且つ「商品ID」が異なる場合のみ「再出品商品」と判断し処理を続行
                                if cellTarget.value == cellComparison.value and cellTarget.offset(0, 1).value != cellComparison.offset(0, 1).value:
                                    #「再出品商品」の「最安値」セルに「ターゲット商品」の「最安値」を設定
                                    cellComparison.offset(0, 2).value = cellTarget.offset(0, 2).value;
                                    # 書式設定 [表示形式：通貨]
                                    cellComparison.offset(0, 2).number_format = f'¥{builtin_format_code(3)};¥-{builtin_format_code(3)}';
                                    
                                    self.m_objLog.LogInfo(f'「再出品商品」の最安値を {cellComparison.offset(0, 2).value}円に設定しました');
                                    self.m_objLog.LogInfo(f'コピー元「商品ID：{cellTarget.offset(0, 1).value} ({cellTarget.row}行目)」→ コピー先「商品ID：{cellComparison.offset(0, 1).value} ({cellComparison.row}行目)」');
                                #「比較商品」が「ターゲット商品」の場合はスキップ
                                else:
                                    continue;
                                
            except Exception as e:
                # 上書き保存
                xlBook.save(strFilePath);
                # 閉じる
                xlBook.close();
                self.m_objLog.LogError(f'「再出品商品」の最安値設定中に不具合が生じたので途中までを保存し終了します');
            
            # 上書き保存
            xlBook.save(strFilePath);
            # 閉じる
            xlBook.close();
            
        except FileNotFoundError as e:
            self.m_objLog.LogCritical(f'{self.m_objSettingFile.FILE_PRODUCT_LIST}の取得に失敗しました');