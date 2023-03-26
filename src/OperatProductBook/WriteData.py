from typing import List, Tuple;
from openpyxl.styles.numbers import builtin_format_code;
import os;
import openpyxl;
import Log;
import Common.Common;
import ExhibitProductInfo;
import Common.ProductBook.Define.PositionPriceCut;
import Common.ProductBook.Define.PositionExhibit;
import Common.ProductBook.Settings;
import Common.ProductBook.Define;
import Common.ProductBook;
import ProductInfo;


class WriteData:
    """
     @class     書き込みクラス
    """
    # ログクラス
    LOG: Log.Log = Log.Log();


    # 共通設定クラス
    COMMON: Common.Common.Common = Common.Common.Common.staticCommon();


# 商品ブックの定数クラス
    EX_SETTING_PRODUCT_BOOK: Common.ProductBook.Settings.Settings = Common.ProductBook.Settings.Settings.staticSettings();
    

    # 値下げ商品のセル位置
    EX_CELL_POSITION_PRICE_CUT: Common.ProductBook.Define.PositionPriceCut.PositionPriceCut = Common.ProductBook.Define.PositionPriceCut.PositionPriceCut.staticPositionPriceCut();
    

    # 出品商品のセル位置
    EX_CELL_POSITION_EXHIBIT: Common.ProductBook.Define.PositionExhibit.PositionExhibit = Common.ProductBook.Define.PositionExhibit.PositionExhibit.staticPositionExhibit();
    

    # 商品ブックパス
    PATH_PRODUCT_BOOK: str = os.path.normpath(os.path.join(COMMON.EXE_BASE_FILE_PATH, f'..\\doc\\{EX_SETTING_PRODUCT_BOOK.FILE_PRODUCT_LIST}'));



    def writeGetProductInfo(self, lstWriteProductInfo: List) ->  None:
        """
         @details       値下げ商品取得
         @param[in]     lstWriteProductInfo     値下げ商品情報
        """
        # データファイルに記録されてある商品情報と出品中の商品情報に差異がない場合、追加記録せず終了
        if len(lstWriteProductInfo) == 0:
            self.LOG.LogInfo('データファイルと出品中の商品情報に差異がないため追加記録は行いません');
            return;
        
        try:
            xlBook          = openpyxl.load_workbook(self.PATH_PRODUCT_BOOK);
            xlSheet         = xlBook[self.EX_SETTING_PRODUCT_BOOK.SHEET_PRICE_CUT];

            xlMaxRow        : int   = xlSheet.max_row;      # 全範囲での最終行
            xlLastRow       : int   = 1;                    # 必要な商品情報が記載されている範囲での最終行
            nProductCount   : int   = 1;                    # 商品数のインクリメント変数

            # 最終行から逆ループ
            for row in reversed(range(1, xlMaxRow)):
                # 指定列の行が空でなくなるまで遡り最終的な指定列の最終行を探す (※指定列：商品ID)
                if xlSheet.cell(row=row, column=self.EX_CELL_POSITION_PRICE_CUT.PRODUCT_ID + 1).value != None:
                    # 指定列の最終行を取得
                    xlLastRow = row + 1;
                    break;
                    
            try:
                # 最終行から新規商品情報を記載
                for objWriteProductInfo in lstWriteProductInfo:
                    xlSheet.cell(xlLastRow, self.EX_CELL_POSITION_PRICE_CUT.PRODUCT_NAME + 1, value=objWriteProductInfo.getProductName());      # 商品名
                    xlSheet.cell(xlLastRow, self.EX_CELL_POSITION_PRICE_CUT.PRODUCT_ID + 1, value=objWriteProductInfo.getProductId());          # 商品ID
                    
                    self.LOG.LogInfo(f'{nProductCount}つ目の商品情報を記録しました。商品ID［{objWriteProductInfo.getProductId()}］');
                    xlLastRow       += 1;       # 書き込み行のインクリメント
                    nProductCount   += 1;       # 商品番号のインクリメント
                    
            except Exception as e:
                nProductCount += 1;     # 商品番号のインクリメント
                self.LOG.LogError(f'{nProductCount}つ目の商品情報の記録に失敗しました');
                pass;
                
                
            # 上書き保存
            xlBook.save(self.PATH_PRODUCT_BOOK);
            # 閉じる
            xlBook.close();
            
        except Exception as e:
            self.LOG.LogCritical(f'{self.EX_SETTING_PRODUCT_BOOK.FILE_PRODUCT_LIST}の取得に失敗しました');
        
        
        
    def writePriceCutProductStatus(self, lstWriteProductStatus: List) ->  None:
        """
         @details       商品を値下げ後、その時に取得した商品状態(公開停止、削除、売り切れ)を記録
         @param[in]     lstWriteProductStatus   商品状態リスト（公開停止、削除、売り切れ）
        """
        # データファイルの情報と値下げを行った際に得た商品状態に変更がない場合、追加記録せず終了
        if len(lstWriteProductStatus) == 0:
            self.LOG.LogInfo('商品状態の新たな変更が確認できないため追加記録は行いません');
            return;
    
        try:
            strFilePath = os.path.normpath(os.path.join(self.COMMON.EXE_BASE_FILE_PATH, f'..\\doc\\{self.EX_SETTING_PRODUCT_BOOK.FILE_PRODUCT_LIST}'));
            xlBook = openpyxl.load_workbook(self.PATH_PRODUCT_BOOK);
            xlSheet = xlBook[self.EX_SETTING_PRODUCT_BOOK.SHEET_PRICE_CUT];

            try:
                # 範囲データを順次処理
                for tupleTargetCellData in xlSheet.iter_cols(min_row=2, min_col=self.EX_CELL_POSITION_PRICE_CUT.PRODUCT_ID + 1, max_col=self.EX_CELL_POSITION_PRICE_CUT.PRODUCT_ID + 1):
                    # タプルで返ってきた商品IDのデータを分析
                    for xlCellData in tupleTargetCellData:
                        
                        # 商品IDの取得
                        strCellData = xlCellData.value;
                        
                        # 商品状態「公開停止、削除、売り切れ」のいずれかに該当する商品IDが存在するか範囲データ内を全検索
                        for objWriteProductStatus in lstWriteProductStatus:
                            
                            #「公開停止、削除、売り切れ」いずれかに該当がある商品IDがデータ内に存在するか否か
                            if strCellData == objWriteProductStatus.getProductId():
                                
                                # 「備考」セル位置を取得
                                xlCellData.offset(0, 2);
                                
                                if objWriteProductStatus.getProductStopPublishing():
                                    xlCellData.offset(0, 2).value = '公開停止';
                                    self.LOG.LogInfo(f'商品ID：{objWriteProductStatus.getProductId()} の備考欄に「公開停止」を記録しました');
                                    
                                if objWriteProductStatus.getProductDelete():
                                    xlCellData.offset(0, 2).value = '削除';
                                    self.LOG.LogInfo(f'商品ID：{objWriteProductStatus.getProductId()} の備考欄に「削除」を記録しました');
                                    
                                if objWriteProductStatus.getProductSoldOut():
                                    xlCellData.offset(0, 2).value = '売り切れ';
                                    self.LOG.LogInfo(f'商品ID：{objWriteProductStatus.getProductId()} の備考欄に「売り切れ」を記録しました');
                                    
            except Exception as e:
                self.LOG.LogError(f'商品状態の更新に失敗しました');
                pass;
                
            # 上書き保存
            xlBook.save(self.PATH_PRODUCT_BOOK);
            # 閉じる
            xlBook.close();
            self.LOG.LogInfo('全商品状態の更新が完了しました');
                    
        except FileNotFoundError as e:
            self.LOG.LogCritical(f'{self.EX_SETTING_PRODUCT_BOOK.FILE_PRODUCT_LIST}の取得に失敗しました');
            
            

    def writeExhibitProductStatus(self, lstExhibitProductInfo: List) ->  None:
        """
         @details       出品商品を公開できたか否かの結果を記録
         @param[in]     lstExhibitProductInfo   出品に「成功」した商品リスト
        """
        # 商品出品数が「0」の場合は追加記録せず終了
        if len(lstExhibitProductInfo) == 0:
            self.LOG.LogInfo('出品した商品数が「0」なため記録は行いません');
            return;
    
        try:
            xlBook = openpyxl.load_workbook(self.PATH_PRODUCT_BOOK);
            xlSheet = xlBook[self.EX_SETTING_PRODUCT_BOOK.SHEET_DATA];

            try:
                # 範囲データを順次処理
                for tupleTargetCellData in xlSheet.iter_cols(min_row=self.EX_CELL_POSITION_EXHIBIT.GET_DATA_FIRST_ROW, min_col=self.EX_CELL_POSITION_EXHIBIT.PRODUCT_NO + 1, max_col=self.EX_CELL_POSITION_EXHIBIT.PRODUCT_NO + 1):
                    # タプルで返ってきた商品IDのデータを分析
                    for xlCellData in tupleTargetCellData:
                        
                        # 商品No.の取得
                        strCellData = xlCellData.value;
                        
                        # 出品が「成功」した商品No.を順に取り出し比較
                        for objExhibitProductInfo in lstExhibitProductInfo:
                            
                            # 同「商品No.」に"〇"を付与
                            if strCellData == objExhibitProductInfo.getProductNo():
                                
                                # 「出品有無」セル位置を設定 (※対象のセル位置(商品No.)からの相対位置)
                                xlCellData.offset(0, self.EX_CELL_POSITION_EXHIBIT.ALREADY_PRODUCT - 1).value = '〇';
                                self.LOG.LogInfo(f'商品No.[{objExhibitProductInfo.getProductNo()}]に出品完了マークを付与しました');
                                    
            except Exception as e:
                self.LOG.LogError(f'商品状態の更新に失敗しました');
                pass;
                
            # 上書き保存
            xlBook.save(self.PATH_PRODUCT_BOOK);
            # 閉じる
            xlBook.close();
            self.LOG.LogInfo('全商品状態の更新が完了しました');
                    
        except FileNotFoundError as e:
            self.LOG.LogCritical(f'{self.EX_SETTING_PRODUCT_BOOK.FILE_PRODUCT_LIST}の取得に失敗しました');



    def writeResetLowestPrice(self) ->  None:
        """
        @ details       再出品商品の最安値設定
        """
        try:
            xlBook      = openpyxl.load_workbook(self.PATH_PRODUCT_BOOK);
            xlSheet     = xlBook[self.EX_SETTING_PRODUCT_BOOK.SHEET_PRICE_CUT];
            
            try:
                # [ターゲット商品] 指定行~指定列 範囲内のセル値をタプルで取得し 1列単位で処理
                for cols in xlSheet.iter_cols(min_row=2, min_col=1, max_col=1):
                    # [ターゲット商品] 列から各セル値を取り出し処理 (※iter_cols()で範囲を指定しているため「商品名」の 1つしか存在しない)
                    for cellTarget in cols:
                        # [比較商品] 指定行~指定列 範囲内のセル値をタプルで取得し 1列単位で処理
                        for cols in xlSheet.iter_cols(min_row=2, min_col=1, max_col=1):
                            # [比較商品] 列から各セル値を取り出し処理 (※iter_cols()で範囲を指定しているため「商品名」の 1つしか存在しない)
                            for cellComparison in cols:
                                
                                #「ターゲット商品」の「行番号」が「比較商品」の「行番号」より若い場合はスキップ
                                if cellTarget.row >= cellComparison.row:
                                    continue;
                                
                                #「商品名」が同じ且つ「商品ID」が異なる場合のみ「再出品商品」と判断し処理を続行
                                if cellTarget.value == cellComparison.value and cellTarget.offset(0, 1).value != cellComparison.offset(0, 1).value:
                                    #「再出品商品」の「最安値」セルに「ターゲット商品」の「最安値」を設定
                                    cellComparison.offset(0, 2).value = cellTarget.offset(0, 2).value;
                                    # 書式設定 [表示形式：通貨]
                                    cellComparison.offset(0, 2).number_format = f'¥{builtin_format_code(3)};¥-{builtin_format_code(3)}';
                                    
                                    self.LOG.LogInfo(f'「再出品商品」の最安値を {cellComparison.offset(0, 2).value}円に設定しました');
                                    self.LOG.LogInfo(f'コピー元「商品ID：{cellTarget.offset(0, 1).value} ({cellTarget.row}行目)」→ コピー先「商品ID：{cellComparison.offset(0, 1).value} ({cellComparison.row}行目)」');
                                #「比較商品」が「ターゲット商品」の場合はスキップ
                                else:
                                    continue;
                                
            except Exception as e:
                # 上書き保存
                xlBook.save(self.PATH_PRODUCT_BOOK);
                # 閉じる
                xlBook.close();
                self.LOG.LogError(f'「再出品商品」の最安値設定中に不具合が生じたので途中までを保存し終了します');
            
            # 上書き保存
            xlBook.save(self.PATH_PRODUCT_BOOK);
            # 閉じる
            xlBook.close();
            
        except FileNotFoundError as e:
            self.LOG.LogCritical(f'{self.EX_SETTING_PRODUCT_BOOK.FILE_PRODUCT_LIST}の取得に失敗しました');