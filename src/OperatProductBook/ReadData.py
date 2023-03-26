from typing import List, Tuple;
from openpyxl.styles.numbers import builtin_format_code;
import os;
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet;
import Log;
import Common.Common;
import ExhibitProductInfo;
import Common.ProductBook.Define.PositionPriceCut;
import Common.ProductBook.Define.PositionExhibit;
import Common.ProductBook.Settings;
import Common.ProductBook.Define;
import Common.ProductBook;
import ProductInfo;



class ReadData:
    """    
     @class     商品情報の読み込み
    """
    # ログクラス
    LOG: Log.Log = Log.Log();
    

    # 共通設定クラス
    COMMON: Common.Common.Common = Common.Common.Common.staticCommon();
    

    # 商品ブックの定数クラス
    EX_SETTING_PRODUCT_BOOK: Common.ProductBook.Settings.Settings = Common.ProductBook.Settings.Settings.staticSettings();
    

    # 値下げ商品のセル位置
    EX_CELL_POSITION_PRICE_CUT: Common.ProductBook.Define.PositionPriceCut.PositionPriceCut = Common.ProductBook.Define.PositionPriceCut.PositionPriceCut.staticPositionPriceCut();
    

    # 出品商品のセル位置
    EX_CELL_POSITION_EXHIBIT: Common.ProductBook.Define.PositionExhibit.PositionExhibit = Common.ProductBook.Define.PositionExhibit.PositionExhibit.staticPositionExhibit();
    

    # 商品ブックパス
    PATH_PRODUCT_BOOK: str = os.path.normpath(os.path.join(COMMON.EXE_BASE_FILE_PATH, f'..\\doc\\{EX_SETTING_PRODUCT_BOOK.FILE_PRODUCT_LIST}'));
    

    # 出品商品の写真が格納されている親までの絶対パス
    PATH_TO_PHOTO_FOLDER: str = os.path.normpath(os.path.join(COMMON.EXE_BASE_FILE_PATH, f'..\\doc\\{COMMON.FOLDER_PRODUCT_PHOTO}'));

    

    def readPriceCutProduct(self) ->  List[ProductInfo.PriceCutProduct]:
        """
         @detail        値下げ商品リスト取得
         @return        lstPriceCutProduct      値下げ商品情報リスト
        """
        self.LOG.LogInfo(f'［シート名：{self.EX_SETTING_PRODUCT_BOOK.SHEET_PRICE_CUT}］の値下げ商品情報を取得します。');
        
        
        #「値下げ」商品情報リスト
        lstPriceCutProduct: List[ProductInfo.PriceCutProduct] = []
        
        try:
            # 該当の「商品情報」ファイルを取得
            xlBook = openpyxl.load_workbook(self.PATH_PRODUCT_BOOK);
            xlSheet = xlBook[self.EX_SETTING_PRODUCT_BOOK.SHEET_PRICE_CUT];
            
            # 商品数カウンター
            nProductCount: int = 1;
            try:
                for row in xlSheet.iter_rows(min_row=2):
                    # セルの値をリストとして取得
                    xlValues: list = [cell.value for cell in row];
                    
                    # 商品IDが空なら次へ
                    if xlValues[self.EX_CELL_POSITION_PRICE_CUT.PRODUCT_ID] is None:
                        continue;
                    
                    strProductName  : str   = xlValues[self.EX_CELL_POSITION_PRICE_CUT.PRODUCT_NAME];          # 商品名
                    strProductId    : str   = xlValues[self.EX_CELL_POSITION_PRICE_CUT.PRODUCT_ID];            # 商品ID
                    nCheapestPrice  : int   = xlValues[self.EX_CELL_POSITION_PRICE_CUT.CHEAPES_PRICE];    # 最安値
                    
                    # 値下げ商品情報
                    objPriceCutInfo: ProductInfo.PriceCutProduct = ProductInfo.PriceCutProduct(strProductName, strProductId);
                    
                    # 最安値が「空白」でなければ
                    if nCheapestPrice is not None:
                        objPriceCutInfo.setCheapestPrice(nCheapestPrice);
                        
                    # 数値以外であれば警告を出し「0円」で設定
                    else:
                        objPriceCutInfo.setCheapestPrice(0);
                        self.LOG.LogWarning('最安値が「空白」または「文字列」で設定されています');
                        self.LOG.LogWarning(f'最安値を「0円」で設定し直しました。商品ID［{strProductId}］');
                    
                    # 商品情報の追加
                    lstPriceCutProduct.append(objPriceCutInfo);
                    
                    self.LOG.LogInfo(f'{nProductCount}つ目の商品情報の取得に成功しました。商品ID［{strProductId}］');
                    nProductCount += 1;     # 商品番号のインクリメント
                    
            except Exception as e:
                self.LOG.LogError(f'{nProductCount}つ目の商品情報の取得に失敗しました');
                nProductCount += 1;         # 商品番号のインクリメント
                pass;
            
            # 上書き保存
            xlBook.save(self.PATH_PRODUCT_BOOK);
            # 閉じる
            xlBook.close();
            
        except FileNotFoundError as e:
            self.LOG.LogCritical(f'{self.EX_SETTING_PRODUCT_BOOK.FILE_PRODUCT_LIST}の取得に失敗しました');
            raise;
            

        return lstPriceCutProduct;
    
    
    
    def readExhibitProduct(self) ->  List[ExhibitProductInfo.ExhibitProductInfo]:
        """
         @detail        出品商品取得
         @return        lstExhibitProduct   出品商品情報リスト
         @note          ProductList.xlsx(シート名：template, data, master)内の「出品商品」リスト取得
        """
        self.LOG.LogInfo(f'{self.EX_SETTING_PRODUCT_BOOK.FILE_PRODUCT_LIST}内の商品情報を取得します');
        
        # 出品商品のリスト
        lstExhibitProductInfo: List[ExhibitProductInfo.ExhibitProductInfo] = []
        
        try:
            
            # 出品商品情報が入力されている Excelブックを開く
            strExcelBookPath: str = os.path.normpath(os.path.join(self.COMMON.EXE_BASE_FILE_PATH, f'..\\doc\\{self.EX_SETTING_PRODUCT_BOOK.FILE_PRODUCT_LIST}'));
            xlBook: openpyxl.Workbook = openpyxl.load_workbook(strExcelBookPath);
            
            #########################################################
            # シート名：「master」の情報取得
            #########################################################
            self.LOG.LogInfo(f'［シート名：{self.EX_SETTING_PRODUCT_BOOK.SHEET_MASTER}］を取得します');
            
            objExhibitProductCommon: ExhibitProductInfo.ExhibitProductCommon;
            try:
                xlSheetMaster: Worksheet = xlBook[self.EX_SETTING_PRODUCT_BOOK.SHEET_MASTER];
                
                xlValue_ShippingCharge = xlSheetMaster[self.EX_CELL_POSITION_EXHIBIT.SHUPPING_CHARGE].value;    # 配送料の負担
                if xlValue_ShippingCharge == None:
                            xlValue_ShippingCharge = '送料込み（出品者負担）';      #? 記載がなかった時のデフォルト値
                            self.LOG.LogInfo(f'「配送料の負担」項目が未設定のためデフォルト値で設定し直しました。配送料の負担［{xlValue_ShippingCharge}］');
                            
                xlValue_ShippingMethod = xlSheetMaster[self.EX_CELL_POSITION_EXHIBIT.SHUPPING_METHOD].value;    # 配送の方法
                if xlValue_ShippingMethod == None:
                            xlValue_ShippingMethod = '未定';                      #? 記載がなかった時のデフォルト値
                            self.LOG.LogInfo(f'「配送の方法」項目が未設定のためデフォルト値で設定し直しました。配送料の負担［{xlValue_ShippingMethod}］');
                            
                xlValue_ShippingArea = xlSheetMaster[self.EX_CELL_POSITION_EXHIBIT.SHUPPING_AREA].value;        # 配送元の地域
                if xlValue_ShippingArea == None:
                            xlValue_ShippingArea = '東京都'                       #? 記載がなかった時のデフォルト値
                            self.LOG.LogInfo(f'「配送元の地域」項目が未設定のためデフォルト値で設定し直しました。配送料の負担［{xlValue_ShippingArea}］');
                            
                xlValue_ShippingDays = xlSheetMaster[self.EX_CELL_POSITION_EXHIBIT.SHUPPING_DAYS].value;        # 発送までの日数
                if xlValue_ShippingDays == None:
                            xlValue_ShippingDays = '4~7日で発送';                 #? 記載がなかった時のデフォルト値
                            self.LOG.LogInfo(f'「発送までの日数」項目が未設定のためデフォルト値で設定し直しました。配送料の負担［{xlValue_ShippingDays}］');
                
                
                # 出品商品情報クラス
                objExhibitProductCommon = ExhibitProductInfo.ExhibitProductCommon(
                    xlValue_ShippingCharge,
                    xlValue_ShippingMethod,
                    xlValue_ShippingArea,
                    xlValue_ShippingDays);

                
            except Exception as e:
                self.LOG.LogError(f'［シート名：{self.EX_SETTING_PRODUCT_BOOK.SHEET_MASTER}］が存在しません');
                
                # デフォルト値に設定
                objExhibitProductCommon = ExhibitProductInfo.ExhibitProductCommon(
                    '送料込み（出品者負担）',
                    '未定',
                    '東京都',
                    '4~7日で発送'
                );
                
                self.LOG.LogInfo('発送に関する情報をデフォル値で設定し続行します');
                pass;
            
            
            #########################################################
            # シート名：「template」の情報取得
            #########################################################
            self.LOG.LogInfo(f'［シート名：{self.EX_SETTING_PRODUCT_BOOK.SHEET_TEMPLATE}］を取得します');
            xlSheetTemplate = xlBook[self.EX_SETTING_PRODUCT_BOOK.SHEET_TEMPLATE];
            xlValueExplanation  = xlSheetTemplate["B5"];       # 商品説明（テンプレート）
            if xlValueExplanation == '':
                self.LOG.LogWarning(f'シート名「{self.EX_SETTING_PRODUCT_BOOK.SHEET_TEMPLATE}」商品説明の雛形が【空白】のため商品説明の入力欄は空白で設定します');
            
            
            #########################################################
            # シート名：「data」の情報取得
            #########################################################
            self.LOG.LogInfo(f'［シート名：{self.EX_SETTING_PRODUCT_BOOK.SHEET_DATA}］を取得します');
                    
            # インクリメント用変数
            nProductCount           : int = 0;          # 全体商品数
            nProductSuccessCount    : int = 0;          # 取得成功商品数
            nProductSkipCount       : int = 0;          # 取得スキップ商品数
            nProductErrorCount      : int = 0;          # 取得失敗商品数
            
            xlSheetData = xlBook[self.EX_SETTING_PRODUCT_BOOK.SHEET_DATA];
            for row in xlSheetData.iter_rows(min_row=self.EX_CELL_POSITION_EXHIBIT.GET_DATA_FIRST_ROW):
                
                try:    
                    # 商品(全体)数のインクリメント
                    nProductCount += 1;
                    # 商品説明の雛形
                    strExplanation: str = xlValueExplanation.value;

                    # セルの値をリストとして取得
                    xlValues: List = [cell.value for cell in row];
                    
                    # 商品No.が空なら次へ
                    if xlValues[self.EX_CELL_POSITION_EXHIBIT.PRODUCT_NAME] is None:
                        # 商品(スキップ)数のインクリメント
                        self.LOG.LogWarning(f'商品No.[{nProductCount}]の商品が【空白】のためスキップします［シート名：{self.EX_SETTING_PRODUCT_BOOK.SHEET_DATA}］');
                        nProductSkipCount += 1;
                        continue;
                    
                    # 商品が既に出品済みであれば次へ（R列に"〇"が付いていれば）
                    if xlValues[self.EX_CELL_POSITION_EXHIBIT.ALREADY_PRODUCT] == '〇':
                        # 商品(スキップ)数のインクリメント
                        self.LOG.LogWarning(f'商品No.[{nProductCount}]の商品が【既に出品中】のためスキップします［シート名：{self.EX_SETTING_PRODUCT_BOOK.SHEET_DATA}］');
                        nProductSkipCount += 1;
                        continue;
                    

                    # 商品番号
                    strProductNo                : str           = xlValues[self.EX_CELL_POSITION_EXHIBIT.PRODUCT_NO];

                    # 商品写真のフォルダ名
                    strProductPhotoFolderName   : str           = xlValues[self.EX_CELL_POSITION_EXHIBIT.PRODUCT_PHOTO_FOLDER_PATH];

                    # カテゴリー_1
                    strCategory_1               : str           = xlValues[self.EX_CELL_POSITION_EXHIBIT.CATEGORY_1];
                    if strCategory_1            == None:
                        strCategory_1                           = 'その他';                         #? 記載がなかった時のデフォルト値

                    # カテゴリー_2
                    strCategory_2               : str           = xlValues[self.EX_CELL_POSITION_EXHIBIT.CATEGORY_2];
                    if strCategory_2            == None:
                        strCategory_2                           = 'その他';                         #? 記載がなかった時のデフォルト値

                    # カテゴリー_3
                    strCategory_3               : str           = xlValues[self.EX_CELL_POSITION_EXHIBIT.CATEGORY_3];

                    # ブランド_1
                    strBrand_1                  : str           = xlValues[self.EX_CELL_POSITION_EXHIBIT.BRAND_1];

                    # 商品名
                    strProductName              : str           = xlValues[self.EX_CELL_POSITION_EXHIBIT.PRODUCT_NAME];
                    if strProductName           == None:
                        strProductName                          = '商品名記入なし';                  #? 記載がなかった時のデフォルト値

                    # 商品状態
                    strCommodityCondition       : str           = xlValues[self.EX_CELL_POSITION_EXHIBIT.COMMODITY_CONDITION];
                    if strCommodityCondition    == None:
                        strCommodityCondition                   = '新品、未使用';                   #? 記載がなかった場合のデフォルト値

                    # 出品開始額
                    nStartingPrice              : int           = xlValues[self.EX_CELL_POSITION_EXHIBIT.STARTING_PRICE];
                    if nStartingPrice           == None:
                        nStartingPrice                          = 999999;

                    # 商品状態_A
                    strProductDetailsA          : str           = str(xlValues[self.EX_CELL_POSITION_EXHIBIT.PRODUCT_DETAILS_A]);
                    if strProductDetailsA       == None:
                        strProductDetailsA                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_B
                    strProductDetailsB          : str           = str(xlValues[self.EX_CELL_POSITION_EXHIBIT.PRODUCT_DETAILS_B]);
                    if strProductDetailsB       == None:
                        strProductDetailsB                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_C
                    strProductDetailsC          : str           = str(xlValues[self.EX_CELL_POSITION_EXHIBIT.PRODUCT_DETAILS_C]);
                    if strProductDetailsC       == None:
                        strProductDetailsC                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_D
                    strProductDetailsD          : str           = str(xlValues[self.EX_CELL_POSITION_EXHIBIT.PRODUCT_DETAILS_D]);
                    if strProductDetailsD       == None:
                        strProductDetailsD                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_E
                    strProductDetailsE          : str           = str(xlValues[self.EX_CELL_POSITION_EXHIBIT.PRODUCT_DETAILS_E]);
                    if strProductDetailsE       == None:
                        strProductDetailsE                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_F
                    strProductDetailsF          : str           = str(xlValues[self.EX_CELL_POSITION_EXHIBIT.PRODUCT_DETAILS_F]);
                    if strProductDetailsF       == None:
                        strProductDetailsF                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_G
                    strProductDetailsG          : str           = str(xlValues[self.EX_CELL_POSITION_EXHIBIT.PRODUCT_DETAILS_G]);
                    if strProductDetailsG       == None:
                        strProductDetailsG                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_H
                    strProductDetailsH          : str           = str(xlValues[self.EX_CELL_POSITION_EXHIBIT.PRODUCT_DETAILS_H]);
                    if strProductDetailsH       == None:
                        strProductDetailsH                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_I
                    strProductDetailsI          : str           = str(xlValues[self.EX_CELL_POSITION_EXHIBIT.PRODUCT_DETAILS_I]);
                    if strProductDetailsI       == None:
                        strProductDetailsI                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_J
                    strProductDetailsJ          : str           = str(xlValues[self.EX_CELL_POSITION_EXHIBIT.PRODUCT_DETAILS_J]);
                    if strProductDetailsJ       == None:
                        strProductDetailsJ                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_K
                    strProductDetailsK          : str           = str(xlValues[self.EX_CELL_POSITION_EXHIBIT.PRODUCT_DETAILS_K]);
                    if strProductDetailsK       == None:
                        strProductDetailsK                      = ' ';                               #? 記載がなかった場合のデフォルト値

                    # 商品状態_L
                    strProductDetailsL          : str           = str(xlValues[self.EX_CELL_POSITION_EXHIBIT.PRODUCT_DETAILS_L]);
                    if strProductDetailsL       == None:
                        strProductDetailsL                      = ' ';                               #? 記載がなかった場合のデフォルト値
                        
                    
                    
                    #############################################################################
                    # 設定されたデータ(商品写真パス)から商品写真の取得
                    #############################################################################
                    # 商品写真の指定フォルダを親までの絶対パスと結合
                    strTargetPhotoFolderPath: str = self.PATH_TO_PHOTO_FOLDER + f'\\{strProductPhotoFolderName}';
                
                    # フォルダ内全ての商品写真のファイル名を絶対パスで追加      #! (確認用)
                    lstProductPhotoFileName: List = [];
                    for strPhotoFileName in os.listdir(strTargetPhotoFolderPath):
                        lstProductPhotoFileName.append(f'{strTargetPhotoFolderPath}\\{strPhotoFileName}');
                    self.LOG.LogDebug(f'商品No.[{strProductNo}]内、全てのファイル名\n{lstProductPhotoFileName}');       # フォルダ内に存在する商品写真ファイル名を全て表示
                    
                    
                    if len(lstProductPhotoFileName) < 1:
                        self.LOG.LogWarning(f'商品No.[{strProductNo}] の商品写真が 1つも見つかりませんでした');
                    elif len(lstProductPhotoFileName) >= 10:
                        self.LOG.LogWarning(f'商品No.[{strProductNo}]の商品写真が 10個(制限枚数)を超えているため 11個目以降の写真は挿入されません');
                        
                        # 挿入されない商品の警告
                        for strProductPhotoFileName in lstProductPhotoFileName[10::-1]:
                            self.LOG.LogWarning(f'挿入されない商品写真［{strProductPhotoFileName}］');
                            lstProductPhotoFileName.remove(strProductPhotoFileName);
                            
                        self.LOG.LogDebug(f'挿入される写真は以下です［商品No.[{strProductNo}]］\n{lstProductPhotoFileName}');
                        
                    
                    #############################################################################
                    # 設定されたデータ(商品説明の各設定値)から商品説明のテンプレートを作成(整形)
                    #############################################################################
                    # 商品ごとの設定値を「商品説明」の雛形に文字列を置換し整形
                    strExplanation = strExplanation.replace("$A$", strProductDetailsA);
                    strExplanation = strExplanation.replace("$B$", strProductDetailsB);
                    strExplanation = strExplanation.replace("$C$", strProductDetailsC);
                    strExplanation = strExplanation.replace("$D$", strProductDetailsD);
                    strExplanation = strExplanation.replace("$E$", strProductDetailsE);
                    strExplanation = strExplanation.replace("$F$", strProductDetailsF);
                    strExplanation = strExplanation.replace("$G$", strProductDetailsG);
                    strExplanation = strExplanation.replace("$H$", strProductDetailsH);
                    strExplanation = strExplanation.replace("$I$", strProductDetailsI);
                    strExplanation = strExplanation.replace("$J$", strProductDetailsJ);
                    strExplanation = strExplanation.replace("$K$", strProductDetailsK);
                    strExplanation = strExplanation.replace("$L$", strProductDetailsL);
                    self.LOG.LogDebug(f'商品No.[{strProductNo}]の商品説明\n{strExplanation}');
                    
                    
                    
                    # 出品商品情報（シート名：data）登録
                    objExhibitProductInfo: ExhibitProductInfo.ExhibitProductInfo = ExhibitProductInfo.ExhibitProductInfo(
                        strProductNo,                   # 商品番号
                        strTargetPhotoFolderPath,       # 商品写真フォルダまでの絶対パス
                        strCategory_1,                  # カテゴリ_1
                        strCategory_2,                  # カテゴリ_2
                        strCategory_3,                  # カテゴリ_3
                        strBrand_1,                     # ブランド_1
                        strProductName,                 # 商品名
                        strCommodityCondition,          # 商品状態
                        nStartingPrice,                 # 出品価格
                        strExplanation                  # 商品説明
                        );
                    
                    
                    # 共通設定部（シート名：master）登録
                    objExhibitProductInfo.setExhibitProductCommon(objExhibitProductCommon);
                    
                    
                    #############################################################################
                    # 出品商品情報を追加
                    #############################################################################
                    lstExhibitProductInfo.append(objExhibitProductInfo);
                    
                    nProductSuccessCount += 1;      # 商品(成功)数のインクリメント
                    self.LOG.LogInfo(f'商品No.[{objExhibitProductInfo.getProductNo()}]の商品情報の取得に成功しました');
                    
                except Exception as e:
                    self.LOG.LogError(f'商品No.[{nProductCount}]の商品情報の取得に失敗しました');
                    nProductErrorCount += 1;        # 商品(エラー)数のインクリメント
                    continue;
                
            # 商品数読み込み結果の出力
            self.LOG.LogInfo(f'商品全体数：{nProductCount}  成功：{nProductSuccessCount}  スキップ：{nProductSkipCount}  失敗：{nProductErrorCount}');
            
            # 閉じる
            xlBook.close();
            
        except FileNotFoundError as e:
            self.LOG.LogCritical(f'{self.EX_SETTING_PRODUCT_BOOK.FILE_PRODUCT_LIST}の取得に失敗しました');
            return lstExhibitProductInfo;
        
        return lstExhibitProductInfo;
